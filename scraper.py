"""
Full-Site Image Scraper
=======================
Enter a homepage URL → automatically crawls all internal pages →
scrapes images (incl. SVGs) → saves an Excel report with one sheet
per page plus a Summary sheet with accessibility metrics.
"""

import os
import re
import time
import hashlib
import threading
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from io import BytesIO
from urllib.parse import urljoin, urlparse, urlunparse
from collections import deque

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from PIL import Image as PILImage, ImageDraw
import fitz  # PyMuPDF

from selenium import webdriver
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# ── Constants ────────────────────────────────────────────────────────────────

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
}

THUMB_MAX_W = 150
THUMB_MAX_H = 100

# File extensions to skip during crawling (non-HTML resources)
SKIP_EXTENSIONS = {
    ".pdf", ".zip", ".rar", ".7z", ".tar", ".gz",
    ".jpg", ".jpeg", ".png", ".gif", ".webp", ".svg", ".ico", ".bmp",
    ".mp4", ".mp3", ".avi", ".mov", ".wmv", ".flv", ".wav", ".ogg",
    ".doc", ".docx", ".xls", ".xlsx", ".ppt", ".pptx",
    ".css", ".js", ".json", ".xml", ".txt", ".csv",
    ".exe", ".msi", ".dmg", ".apk",
}

# ── Excel Styles ─────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Segoe UI", size=10)
BODY_FONT_BOLD = Font(name="Segoe UI", size=10, bold=True)
URL_FONT = Font(name="Segoe UI", size=9, color="2563EB")
IMG_CELL_FILL = PatternFill(start_color="D9DCE0", end_color="D9DCE0", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
MISSING_FONT = Font(name="Segoe UI", size=10, bold=True, color="DC2626")
MISSING_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
SUMMARY_TITLE_FONT = Font(name="Segoe UI", size=18, bold=True, color="1F2937")
SUMMARY_LABEL_FONT = Font(name="Segoe UI", size=12, bold=True)
SUMMARY_VALUE_FONT = Font(name="Segoe UI", size=14, bold=True, color="6366F1")
SCORE_GOOD = Font(name="Segoe UI", size=14, bold=True, color="16A34A")
SCORE_MED = Font(name="Segoe UI", size=14, bold=True, color="CA8A04")
SCORE_BAD = Font(name="Segoe UI", size=14, bold=True, color="DC2626")
GREEN_VAL = Font(name="Segoe UI", size=14, bold=True, color="16A34A")
RED_VAL = Font(name="Segoe UI", size=14, bold=True, color="DC2626")
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)


# ── Selenium Helpers ─────────────────────────────────────────────────────────

def create_driver() -> webdriver.Chrome:
    opts = ChromeOptions()
    opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--window-size=1920,1080")
    opts.add_argument(f"user-agent={HEADERS['User-Agent']}")
    opts.add_argument("--log-level=3")
    driver = webdriver.Chrome(options=opts)
    driver.set_page_load_timeout(30)
    return driver


def load_page(driver: webdriver.Chrome, url: str):
    """Navigate to URL and wait for content."""
    driver.get(url)
    try:
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
    except Exception:
        pass
    time.sleep(2)


# ── Site Crawler ─────────────────────────────────────────────────────────────

def normalize_url(url: str) -> str:
    """Normalize a URL for deduplication."""
    parsed = urlparse(url)
    # Remove fragment, normalize path
    path = parsed.path.rstrip("/") or "/"
    return urlunparse((
        parsed.scheme, parsed.netloc.lower(), path,
        parsed.params, parsed.query, ""
    ))


def discover_links(driver: webdriver.Chrome, page_url: str, base_domain: str) -> list[str]:
    """Extract all same-domain links from the current page."""
    soup = BeautifulSoup(driver.page_source, "html.parser")
    links = []
    seen = set()

    for a_tag in soup.find_all("a", href=True):
        href = a_tag["href"].strip()
        if not href or href.startswith(("#", "mailto:", "tel:", "javascript:")):
            continue

        abs_url = urljoin(page_url, href)
        parsed = urlparse(abs_url)

        # Same domain only
        if parsed.netloc.lower().replace("www.", "") != base_domain:
            continue

        # Skip non-HTML resources
        ext = os.path.splitext(parsed.path)[1].lower()
        if ext in SKIP_EXTENSIONS:
            continue

        norm = normalize_url(abs_url)
        if norm not in seen:
            seen.add(norm)
            links.append(norm)

    return links


def crawl_site(
    driver: webdriver.Chrome,
    start_url: str,
    max_pages: int,
    on_status=None,
    on_log=None,
) -> list[str]:
    """
    BFS crawl from start_url, collecting up to max_pages internal URLs.
    Returns a list of URLs in discovery order.
    """
    start_norm = normalize_url(start_url)
    parsed = urlparse(start_norm)
    base_domain = parsed.netloc.lower().replace("www.", "")

    visited = set()
    queue = deque([start_norm])
    ordered = []

    while queue and len(ordered) < max_pages:
        url = queue.popleft()
        if url in visited:
            continue
        visited.add(url)
        ordered.append(url)

        if on_status:
            on_status(f"🌐 Crawling ({len(ordered)}/{max_pages}): discovering pages…")
        if on_log:
            on_log(f"  🔗 [{len(ordered)}] {url}")

        try:
            load_page(driver, url)
            new_links = discover_links(driver, url, base_domain)
            for link in new_links:
                if link not in visited:
                    queue.append(link)
        except Exception as e:
            if on_log:
                on_log(f"    ⚠ Could not load: {e}")

    return ordered


# ── Image Extraction ─────────────────────────────────────────────────────────

def extract_images(driver: webdriver.Chrome, url: str) -> list[dict]:
    """Extract only visible, meaningful images from the currently loaded page."""
    images = []
    seen = set()

    # Use Selenium to find visible <img> elements
    img_elements = driver.find_elements(By.TAG_NAME, "img")
    for el in img_elements:
        try:
            # Skip hidden / zero-size images
            if not el.is_displayed():
                continue
            w = el.size.get("width", 0)
            h = el.size.get("height", 0)
            # Skip tracking pixels and tiny decorative images
            if w < 5 or h < 5:
                continue
        except Exception:
            continue

        src = (
            el.get_attribute("src") or ""
        ).strip()
        if not src:
            src = (el.get_attribute("data-src") or "").strip()
        if not src:
            src = (el.get_attribute("data-lazy-src") or "").strip()
        if not src:
            srcset = (el.get_attribute("srcset") or "").strip()
            if srcset:
                src = srcset.split(",")[0].strip().split(" ")[0]
        if not src or src.startswith("data:"):
            continue

        abs_url = urljoin(url, src)
        if abs_url in seen:
            continue
        seen.add(abs_url)

        raw_alt = el.get_attribute("alt")
        alt = raw_alt.strip() if raw_alt is not None else None

        images.append({"src": abs_url, "alt": alt})

    # Inline SVGs — only visible ones with meaningful content
    svg_elements = driver.find_elements(By.TAG_NAME, "svg")
    for i, el in enumerate(svg_elements):
        try:
            if not el.is_displayed():
                continue
            w = el.size.get("width", 0)
            h = el.size.get("height", 0)
            if w < 10 or h < 10:
                continue
        except Exception:
            continue

        try:
            svg_html = el.get_attribute("outerHTML") or ""
        except Exception:
            continue

        if len(svg_html) < 50:
            continue

        key = f"__inline_svg_{i}__"
        if key in seen:
            continue
        seen.add(key)

        images.append({
            "src": f"[inline SVG #{i+1}]",
            "alt": el.get_attribute("aria-label") or el.get_attribute("title") or None,
            "svg_content": svg_html,
        })

    return images


# ── Image Download & Processing ──────────────────────────────────────────────

def download_image(url: str, save_dir: str) -> tuple[str | None, str | None]:
    """Download image. Returns (local_path, svg_source_code_or_None)."""
    try:
        resp = requests.get(url, headers=HEADERS, timeout=15, stream=True)
        resp.raise_for_status()

        parsed = urlparse(url)
        basename = os.path.basename(parsed.path) or "image"
        basename = re.sub(r"[?#].*", "", basename)
        basename = re.sub(r'[<>:"/\\|?*]', '_', basename)

        if "." not in basename:
            ct = resp.headers.get("Content-Type", "")
            ext = ".jpg"
            if "png" in ct: ext = ".png"
            elif "gif" in ct: ext = ".gif"
            elif "webp" in ct: ext = ".webp"
            elif "svg" in ct: ext = ".svg"
            basename += ext

        save_path = os.path.join(save_dir, basename)
        counter = 1
        name, ext = os.path.splitext(save_path)
        while os.path.exists(save_path):
            save_path = f"{name}_{counter}{ext}"
            counter += 1

        with open(save_path, "wb") as f:
            for chunk in resp.iter_content(8192):
                f.write(chunk)

        svg_source = None
        if save_path.lower().endswith(".svg"):
            # Read SVG source before conversion (for color analysis later)
            try:
                with open(save_path, "r", encoding="utf-8", errors="ignore") as sf:
                    svg_source = sf.read()
            except Exception:
                pass
            save_path = convert_svg_to_png(save_path)
        if save_path and save_path.lower().endswith(".webp"):
            save_path = convert_webp_to_png(save_path)

        return save_path, svg_source
    except Exception as e:
        print(f"[Download] Failed: {url[:80]} \u2013 {e}")
        return None, None


def save_inline_svg(svg_content: str, save_dir: str, index: int) -> str | None:
    try:
        svg_path = os.path.join(save_dir, f"inline_svg_{index}.svg")
        with open(svg_path, "w", encoding="utf-8") as f:
            f.write(svg_content)
        return convert_svg_to_png(svg_path)
    except Exception as e:
        print(f"[InlineSVG] Failed: {e}")
        return None


def convert_svg_to_png(svg_path: str) -> str | None:
    try:
        png_path = os.path.splitext(svg_path)[0] + ".png"
        doc = fitz.open(svg_path)
        page = doc[0]
        pix = page.get_pixmap(matrix=fitz.Matrix(2.0, 2.0), alpha=True)
        doc.close()
        img = PILImage.frombytes("RGBA", [pix.width, pix.height], pix.samples)
        img.save(png_path, "PNG")
        try:
            os.remove(svg_path)
        except OSError:
            pass
        return png_path
    except Exception as e:
        print(f"[SVG→PNG] Failed: {svg_path}: {e}")
        return svg_path


def convert_webp_to_png(webp_path: str) -> str | None:
    try:
        png_path = os.path.splitext(webp_path)[0] + ".png"
        PILImage.open(webp_path).save(png_path, "PNG")
        try:
            os.remove(webp_path)
        except OSError:
            pass
        return png_path
    except Exception as e:
        print(f"[WebP→PNG] Failed: {e}")
        return webp_path


# ── SVG Color Analysis ────────────────────────────────────────────────────────

_CSS_COLORS = {
    "black": (0,0,0), "white": (255,255,255), "red": (255,0,0),
    "green": (0,128,0), "blue": (0,0,255), "yellow": (255,255,0),
    "cyan": (0,255,255), "magenta": (255,0,255), "gray": (128,128,128),
    "grey": (128,128,128), "silver": (192,192,192), "maroon": (128,0,0),
    "olive": (128,128,0), "lime": (0,255,0), "aqua": (0,255,255),
    "teal": (0,128,128), "navy": (0,0,128), "fuchsia": (255,0,255),
    "purple": (128,0,128), "orange": (255,165,0), "pink": (255,192,203),
    "brown": (165,42,42), "coral": (255,127,80), "crimson": (220,20,60),
    "darkblue": (0,0,139), "darkgreen": (0,100,0), "darkred": (139,0,0),
    "gold": (255,215,0), "indigo": (75,0,130), "ivory": (255,255,240),
    "khaki": (240,230,140), "lavender": (230,230,250), "linen": (250,240,230),
    "peru": (205,133,63), "plum": (221,160,221), "salmon": (250,128,114),
    "sienna": (160,82,45), "skyblue": (135,206,235), "tan": (210,180,140),
    "tomato": (255,99,71), "violet": (238,130,238), "wheat": (245,222,179),
    "none": None, "transparent": None, "currentcolor": None,
}


def _parse_color(raw: str) -> tuple[int, int, int] | None:
    """Convert a CSS color string to (R, G, B)."""
    s = raw.strip().lower()
    if not s:
        return None
    if s in _CSS_COLORS:
        return _CSS_COLORS[s]
    m = re.match(r"^#([0-9a-f]{3,8})$", s)
    if m:
        h = m.group(1)
        if len(h) == 3:
            return (int(h[0]*2, 16), int(h[1]*2, 16), int(h[2]*2, 16))
        if len(h) >= 6:
            return (int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16))
    m = re.match(r"rgba?\(\s*(\d+)\s*,\s*(\d+)\s*,\s*(\d+)", s)
    if m:
        return (int(m.group(1)), int(m.group(2)), int(m.group(3)))
    return None


def _extract_svg_colors(svg_code: str) -> list[tuple[int, int, int]]:
    """Extract fill/stroke colors from SVG source."""
    colors = []
    for attr in re.finditer(r'(?:fill|stroke)\s*=\s*"([^"]+)"', svg_code):
        c = _parse_color(attr.group(1))
        if c:
            colors.append(c)
    for style in re.finditer(r'style\s*=\s*"([^"]+)"', svg_code):
        for prop in re.finditer(r'(?:fill|stroke|color)\s*:\s*([^;]+)', style.group(1)):
            c = _parse_color(prop.group(1))
            if c:
                colors.append(c)
    for css in re.finditer(r'<style[^>]*>(.*?)</style>', svg_code, re.DOTALL):
        for prop in re.finditer(r'(?:fill|stroke|color)\s*:\s*([^;}\s]+)', css.group(1)):
            c = _parse_color(prop.group(1))
            if c:
                colors.append(c)
    return colors


def _luminance(r: int, g: int, b: int) -> float:
    """Perceived luminance 0..1."""
    return (0.299 * r + 0.587 * g + 0.114 * b) / 255.0


def _svg_contrast_bg(svg_code: str) -> tuple[int, int, int]:
    """Analyze SVG colors and return a contrasting background."""
    colors = _extract_svg_colors(svg_code)
    if not colors:
        return (220, 220, 220)
    avg_lum = sum(_luminance(*c) for c in colors) / len(colors)
    if avg_lum > 0.7:      # light/white SVG  → dark bg
        return (50, 55, 65)
    elif avg_lum < 0.3:    # dark SVG         → light bg
        return (240, 242, 245)
    else:                  # mid-tone         → neutral
        return (230, 232, 235)


def _is_light_or_transparent(img: PILImage.Image) -> bool:
    """Check if image is mostly white/light or transparent."""
    rgba = img.convert("RGBA")
    data = list(rgba.getdata())
    if not data:
        return False
    total = len(data)
    light_count = sum(1 for r, g, b, a in data if a < 128 or (r > 220 and g > 220 and b > 220))
    return light_count / total > 0.4


def make_thumbnail(image_path: str, bg_hint: tuple[int, int, int] | None = None) -> str | None:
    """
    Create a thumbnail with the right background for visibility.
    bg_hint: from SVG color analysis — used instead of auto-detection.
    """
    try:
        img = PILImage.open(image_path)
        if img.mode == "P":
            img = img.convert("RGBA")
        elif img.mode not in ("RGBA",):
            img = img.convert("RGBA")

        img.thumbnail((THUMB_MAX_W, THUMB_MAX_H), PILImage.LANCZOS)

        pad = 8
        bg_w = img.width + pad * 2
        bg_h = img.height + pad * 2

        if bg_hint:
            background = PILImage.new("RGBA", (bg_w, bg_h), bg_hint + (255,))
        elif _is_light_or_transparent(img):
            background = PILImage.new("RGBA", (bg_w, bg_h), (50, 55, 65, 255))
        else:
            background = PILImage.new("RGBA", (bg_w, bg_h), (230, 232, 235, 255))

        background.paste(img, (pad, pad), mask=img)
        final = background.convert("RGB")

        border_color = (200, 200, 200) if bg_hint and _luminance(*bg_hint) < 0.4 else (140, 140, 140)
        draw = ImageDraw.Draw(final)
        draw.rectangle([0, 0, bg_w - 1, bg_h - 1], outline=border_color, width=2)

        thumb_path = image_path + "_thumb.png"
        final.save(thumb_path, "PNG")
        return thumb_path
    except Exception as e:
        print(f"[Thumbnail] Failed: {image_path}: {e}")
        return None


# ── Excel Export ─────────────────────────────────────────────────────────────

def _safe_sheet_name(url: str, existing_names: set) -> str:
    """Generate a valid Excel sheet name (≤31 chars, unique) from a URL path."""
    parsed = urlparse(url)
    path = parsed.path.strip("/")
    if not path:
        name = "Homepage"
    else:
        # Replace slashes with underscores, keep it short
        name = path.replace("/", "_")
        # Remove special chars
        name = re.sub(r'[\\/*?:\[\]]', '', name)

    # Truncate to 28 chars to leave room for dedup suffix
    name = name[:28]

    # Deduplicate
    base = name
    counter = 2
    while name in existing_names:
        name = f"{base[:25]}_{counter}"
        counter += 1

    existing_names.add(name)
    return name


def save_to_excel(pages_data: list[dict], output_path: str, on_progress=None):
    """
    pages_data: list of {'url': str, 'images': [{'src','alt','local_path'}, ...]}
    """
    wb = Workbook()

    # ── Summary Sheet ──
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _build_summary(ws_summary, pages_data)

    # ── Per-page sheets ──
    sheet_names = {"Summary"}
    total_images = sum(len(p["images"]) for p in pages_data)
    processed = 0

    for page in pages_data:
        images = page.get("images", [])
        if not images:
            continue  # skip pages with no images
        sheet_name = _safe_sheet_name(page["url"], sheet_names)
        ws = wb.create_sheet(sheet_name)
        _build_page_sheet(ws, page["url"], images)

        # Insert thumbnails
        for idx, item in enumerate(images):
            row = idx + 3  # row 1=page URL, row 2=headers, data starts row 3
            local_path = item.get("local_path")
            cell_b = ws.cell(row=row, column=2)

            if local_path and os.path.isfile(local_path):
                # SVG color-aware background
                bg_hint = None
                svg_code = item.get("svg_content")
                if svg_code:
                    bg_hint = _svg_contrast_bg(svg_code)
                thumb = make_thumbnail(local_path, bg_hint=bg_hint)
                if thumb:
                    try:
                        xl_img = XlImage(thumb)
                        xl_img.anchor = f"B{row}"
                        ws.add_image(xl_img)
                    except Exception:
                        cell_b.value = "(image error)"
                else:
                    cell_b.value = "(unsupported)"
            else:
                cell_b.value = "(download failed)"

            processed += 1
            if on_progress:
                on_progress(processed, total_images)

    wb.save(output_path)


def _build_summary(ws, pages_data: list[dict]):
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 25

    # Title
    ws.merge_cells("B2:C2")
    c = ws.cell(row=2, column=2, value="📊  Image Scraper Report")
    c.font = SUMMARY_TITLE_FONT
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[2].height = 40

    # Metrics
    total = sum(len(p["images"]) for p in pages_data)
    with_alt = sum(
        1 for p in pages_data for img in p["images"]
        if img.get("alt") is not None and img["alt"].strip()
    )
    missing = total - with_alt
    score = (with_alt / total * 100) if total > 0 else 0

    metrics = [
        ("Total Images Scanned", str(total), SUMMARY_VALUE_FONT),
        ("Images WITH Alt Text", str(with_alt), GREEN_VAL),
        ("Images MISSING Alt Text", str(missing), RED_VAL),
        ("Accessibility Score", f"{score:.1f}%",
         SCORE_GOOD if score >= 80 else SCORE_MED if score >= 50 else SCORE_BAD),
    ]

    for i, (label, value, font) in enumerate(metrics):
        row = 4 + i
        ws.row_dimensions[row].height = 32

        lc = ws.cell(row=row, column=2, value=label)
        lc.font = SUMMARY_LABEL_FONT
        lc.alignment = Alignment(vertical="center")
        lc.border = THIN_BORDER

        vc = ws.cell(row=row, column=3, value=value)
        vc.font = font
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.border = THIN_BORDER

    # Pages + counts
    row = 4 + len(metrics) + 1
    ws.row_dimensions[row].height = 30
    ws.cell(row=row, column=2, value="Pages Crawled").font = Font(
        name="Segoe UI", size=13, bold=True, color="1F2937"
    )

    row += 1
    for ci, h in enumerate(["Page URL", "Images", "Sheet"], 2):
        c = ws.cell(row=row, column=ci, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER
    ws.column_dimensions["D"].width = 22
    ws.row_dimensions[row].height = 28

    sheet_names_set = {"Summary"}
    for i, page in enumerate(pages_data):
        r = row + 1 + i
        sn = _safe_sheet_name(page["url"], sheet_names_set)

        uc = ws.cell(row=r, column=2, value=page["url"])
        uc.font = URL_FONT
        uc.alignment = Alignment(vertical="center", wrap_text=True)
        uc.border = THIN_BORDER

        cc = ws.cell(row=r, column=3, value=len(page["images"]))
        cc.font = BODY_FONT_BOLD
        cc.alignment = Alignment(horizontal="center", vertical="center")
        cc.border = THIN_BORDER

        sc = ws.cell(row=r, column=4, value=sn)
        sc.font = Font(name="Segoe UI", size=9, color="6366F1")
        sc.alignment = Alignment(horizontal="center", vertical="center")
        sc.border = THIN_BORDER

        ws.row_dimensions[r].height = 24


def _build_page_sheet(ws, page_url: str, images: list[dict]):
    """Build a per-page sheet with image data."""
    # Page URL in row 1
    ws.merge_cells("A1:D1")
    c = ws.cell(row=1, column=1, value=f"📄 {page_url}")
    c.font = Font(name="Segoe UI", size=11, bold=True, color="1E40AF")
    c.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    c.alignment = Alignment(vertical="center")
    c.border = THIN_BORDER
    ws.row_dimensions[1].height = 32

    # Column widths
    ws.column_dimensions["A"].width = 8    # S.No
    ws.column_dimensions["B"].width = 24   # Image Preview
    ws.column_dimensions["C"].width = 55   # Image Source URL
    ws.column_dimensions["D"].width = 40   # Alt Text

    # Reassign row 1 as the URL bar; headers go to row 2
    # Actually, let's put headers right after the URL bar
    # Row 1 = Page URL, Row 2 = headers, Row 3+ = data

    # Headers in row 2 (we used row 1 for page URL)
    # Wait, we already used merge on row 1. Let me restructure:
    # Row 1 = merged page URL bar
    # Row 2 = column headers
    # Row 3+ = data

    headers = ["S.No", "Image Preview", "Image Source URL", "Alt Text"]
    for ci, h in enumerate(headers, 1):
        hc = ws.cell(row=2, column=ci, value=h)
        hc.fill = HEADER_FILL
        hc.font = HEADER_FONT
        hc.alignment = Alignment(horizontal="center", vertical="center")
        hc.border = THIN_BORDER
    ws.row_dimensions[2].height = 28

    for idx, item in enumerate(images):
        row = idx + 3  # data starts at row 3
        ws.row_dimensions[row].height = 90

        fill = ALT_ROW_FILL if idx % 2 == 0 else PatternFill()

        # S.No
        sc = ws.cell(row=row, column=1, value=idx + 1)
        sc.font = BODY_FONT
        sc.alignment = Alignment(horizontal="center", vertical="center")
        sc.fill = fill
        sc.border = THIN_BORDER

        # Image Preview (thumbnail inserted later by save_to_excel)
        bc = ws.cell(row=row, column=2)
        bc.fill = IMG_CELL_FILL
        bc.border = THIN_BORDER
        bc.alignment = Alignment(horizontal="center", vertical="center")

        # Image Source URL
        cc = ws.cell(row=row, column=3, value=item.get("src", ""))
        cc.font = URL_FONT
        cc.alignment = Alignment(vertical="center", wrap_text=True)
        cc.fill = fill
        cc.border = THIN_BORDER

        # Alt Text
        alt = item.get("alt")
        if alt is not None and alt.strip():
            dc = ws.cell(row=row, column=4, value=alt)
            dc.font = BODY_FONT
            dc.fill = fill
        else:
            dc = ws.cell(row=row, column=4, value="⚠ MISSING")
            dc.font = MISSING_FONT
            dc.fill = MISSING_FILL
        dc.alignment = Alignment(vertical="center", wrap_text=True)
        dc.border = THIN_BORDER


# ── Tkinter GUI ──────────────────────────────────────────────────────────────

class ImageScraperApp:
    BG = "#0F172A"
    CARD_BG = "#1E293B"
    ACCENT = "#6366F1"
    ACCENT_HOVER = "#818CF8"
    TEXT = "#F8FAFC"
    SUBTEXT = "#94A3B8"

    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("🖼️ Full-Site Image Scraper")
        self.root.geometry("760x660")
        self.root.configure(bg=self.BG)
        self.root.resizable(False, False)
        self._build_ui()

    def _build_ui(self):
        # Title
        tk.Label(
            self.root, text="Full-Site Image Scraper",
            font=("Segoe UI", 22, "bold"), bg=self.BG, fg=self.TEXT,
        ).pack(pady=(20, 2))
        tk.Label(
            self.root,
            text="Enter a homepage URL → crawls all pages → Excel report",
            font=("Segoe UI", 11), bg=self.BG, fg=self.SUBTEXT,
        ).pack(pady=(0, 14))

        # ── Input card ──
        card = tk.Frame(self.root, bg=self.CARD_BG,
                        highlightbackground="#334155", highlightthickness=1)
        card.pack(padx=28, fill="x")

        tk.Label(
            card, text="Homepage URL", font=("Segoe UI", 10, "bold"),
            bg=self.CARD_BG, fg=self.SUBTEXT,
        ).pack(anchor="w", padx=14, pady=(12, 4))

        self.url_var = tk.StringVar()
        entry = tk.Entry(
            card, textvariable=self.url_var, font=("Segoe UI", 12),
            bg="#0F172A", fg=self.TEXT, insertbackground=self.TEXT,
            relief="flat", highlightthickness=1, highlightbackground="#475569",
            highlightcolor=self.ACCENT,
        )
        entry.pack(padx=14, fill="x", ipady=8)
        entry.bind("<Return>", lambda e: self._start_scrape())

        # Max pages row
        opt_frame = tk.Frame(card, bg=self.CARD_BG)
        opt_frame.pack(padx=14, pady=(10, 0), fill="x")

        tk.Label(
            opt_frame, text="Max Pages to Crawl:",
            font=("Segoe UI", 10), bg=self.CARD_BG, fg=self.SUBTEXT,
        ).pack(side="left")

        self.max_pages_var = tk.IntVar(value=20)
        spinner = tk.Spinbox(
            opt_frame, from_=1, to=100, textvariable=self.max_pages_var,
            font=("Segoe UI", 11), width=5, bg="#0F172A", fg=self.TEXT,
            buttonbackground="#334155", relief="flat",
            highlightthickness=1, highlightbackground="#475569",
        )
        spinner.pack(side="left", padx=(8, 0))

        # Buttons
        btn_frame = tk.Frame(card, bg=self.CARD_BG)
        btn_frame.pack(padx=14, pady=12, fill="x")

        self.scrape_btn = tk.Button(
            btn_frame, text="🚀  Start Crawl & Scrape",
            font=("Segoe UI", 11, "bold"),
            bg=self.ACCENT, fg="white", activebackground=self.ACCENT_HOVER,
            activeforeground="white", relief="flat", cursor="hand2",
            command=self._start_scrape, padx=18, pady=8,
        )
        self.scrape_btn.pack(side="left")

        tk.Button(
            btn_frame, text="📁  Save Location", font=("Segoe UI", 10),
            bg="#334155", fg=self.SUBTEXT, activebackground="#475569",
            activeforeground=self.TEXT, relief="flat", cursor="hand2",
            command=self._choose_save, padx=12, pady=8,
        ).pack(side="right")

        self.save_path_var = tk.StringVar(
            value=os.path.join(os.getcwd(), "scraped_images.xlsx")
        )
        tk.Label(
            card, textvariable=self.save_path_var, font=("Segoe UI", 9),
            bg=self.CARD_BG, fg=self.SUBTEXT, anchor="w",
        ).pack(padx=14, pady=(0, 10), fill="x")

        # ── Progress ──
        pf = tk.Frame(self.root, bg=self.BG)
        pf.pack(padx=28, fill="x", pady=(12, 0))

        self.status_var = tk.StringVar(value="Ready — enter a homepage URL to begin.")
        tk.Label(
            pf, textvariable=self.status_var, font=("Segoe UI", 10),
            bg=self.BG, fg=self.SUBTEXT, anchor="w",
        ).pack(fill="x")

        style = ttk.Style()
        style.theme_use("clam")
        style.configure(
            "Custom.Horizontal.TProgressbar",
            troughcolor="#1E293B", background=self.ACCENT,
            bordercolor=self.BG, lightcolor=self.ACCENT, darkcolor=self.ACCENT,
        )
        self.progress = ttk.Progressbar(
            pf, style="Custom.Horizontal.TProgressbar",
            orient="horizontal", mode="determinate",
        )
        self.progress.pack(fill="x", pady=(6, 0))

        # ── Log ──
        lf = tk.Frame(self.root, bg=self.BG)
        lf.pack(padx=28, pady=(12, 18), fill="both", expand=True)

        self.log_text = tk.Text(
            lf, font=("Consolas", 9), bg="#0F172A", fg="#CBD5E1",
            relief="flat", highlightthickness=1, highlightbackground="#334155",
            insertbackground=self.TEXT, state="disabled", wrap="word",
        )
        sb = ttk.Scrollbar(lf, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        self.log_text.pack(side="left", fill="both", expand=True)

    # ── Helpers ──

    def _log(self, msg: str):
        self.log_text.configure(state="normal")
        self.log_text.insert("end", msg + "\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")

    def _status(self, msg: str):
        self.status_var.set(msg)

    def _choose_save(self):
        p = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="scraped_images.xlsx",
        )
        if p:
            self.save_path_var.set(p)

    # ── Main workflow ──

    def _start_scrape(self):
        url = self.url_var.get().strip()
        if not url:
            messagebox.showwarning("No URL", "Please enter a homepage URL.")
            return
        if not url.startswith(("http://", "https://")):
            url = "https://" + url
            self.url_var.set(url)

        self.scrape_btn.configure(state="disabled", text="⏳  Crawling…")
        self.progress["value"] = 0
        self.log_text.configure(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.configure(state="disabled")

        threading.Thread(
            target=self._worker, args=(url,), daemon=True
        ).start()

    def _worker(self, start_url: str):
        try:
            max_pages = self.max_pages_var.get()
            save_xlsx = self.save_path_var.get()
            dl_dir = os.path.join(os.path.dirname(save_xlsx), "downloaded_images")
            os.makedirs(dl_dir, exist_ok=True)

            # ── Phase 1: Crawl ──
            self._status("🚀 Starting browser…")
            self._log("→ Launching headless Chrome…")
            driver = create_driver()
            self._log("✓ Browser ready.\n")

            self._log(f"━━━ Phase 1: Crawling site (max {max_pages} pages) ━━━")
            page_urls = crawl_site(
                driver, start_url, max_pages,
                on_status=self._status, on_log=self._log,
            )
            self._log(f"\n✓ Discovered {len(page_urls)} page(s).\n")

            # ── Phase 2: Scrape images from each page ──
            self._log("━━━ Phase 2: Scraping images ━━━")
            pages_data = []
            inline_svg_counter = 0
            seen_svg_hashes = set()  # Deduplicate SVGs across all pages

            for pi, url in enumerate(page_urls):
                self._status(f"🔍 Scraping page {pi+1}/{len(page_urls)}…")
                self._log(f"\n📄 Page {pi+1}/{len(page_urls)}: {url}")

                try:
                    load_page(driver, url)
                    images = extract_images(driver, url)
                except Exception as e:
                    self._log(f"  ✗ Failed: {e}")
                    pages_data.append({"url": url, "images": []})
                    continue

                # Deduplicate SVGs across pages by content hash
                filtered = []
                for img in images:
                    if "svg_content" in img:
                        svg_hash = hashlib.md5(img["svg_content"].encode()).hexdigest()
                        if svg_hash in seen_svg_hashes:
                            continue  # skip duplicate SVG
                        seen_svg_hashes.add(svg_hash)
                    filtered.append(img)
                images = filtered

                self._log(f"  Found {len(images)} image(s)")

                # Download
                self.progress["maximum"] = max(1, len(images))
                self.progress["value"] = 0

                for i, img in enumerate(images):
                    if "svg_content" in img:
                        inline_svg_counter += 1
                        self._log(f"    🎨 [{i+1}] Converting inline SVG…")
                        local = save_inline_svg(img["svg_content"], dl_dir, inline_svg_counter)
                    else:
                        self._log(f"    ⬇ [{i+1}] {img['src'][:65]}…")
                        local, svg_source = download_image(img["src"], dl_dir)
                        if svg_source:
                            img["svg_content"] = svg_source
                    img["local_path"] = local
                    self.progress["value"] = i + 1

                ok = sum(1 for x in images if x.get("local_path"))
                self._log(f"  ✓ {ok}/{len(images)} downloaded")
                pages_data.append({"url": url, "images": images})

            driver.quit()
            self._log("\n→ Browser closed.")

            # ── Phase 3: Write Excel ──
            total = sum(len(p["images"]) for p in pages_data)
            self._status(f"📊 Writing Excel ({total} images, {len(pages_data)} sheets)…")
            self._log(f"\n━━━ Phase 3: Generating Excel ━━━")
            self._log(f"→ {save_xlsx}")

            save_to_excel(pages_data, save_xlsx, on_progress=self._update_progress)

            # Summary log
            with_alt = sum(
                1 for p in pages_data for img in p["images"]
                if img.get("alt") is not None and img["alt"].strip()
            )
            score = (with_alt / total * 100) if total > 0 else 0

            self._log(f"✓ Excel saved!\n")
            self._log(f"📊 Summary:")
            self._log(f"   Pages crawled:  {len(pages_data)}")
            self._log(f"   Total images:   {total}")
            self._log(f"   With alt text:  {with_alt}")
            self._log(f"   Missing alt:    {total - with_alt}")
            self._log(f"   Accessibility:  {score:.1f}%")
            self._log(f"\n🎉 Done!")

            self._status(
                f"✅ Done — {total} images from {len(pages_data)} pages → Excel!"
            )

        except Exception as e:
            self._status("❌ Error occurred")
            self._log(f"✗ Unexpected error: {e}")
            import traceback
            self._log(traceback.format_exc())
        finally:
            self.root.after(0, lambda: self.scrape_btn.configure(
                state="normal", text="🚀  Start Crawl & Scrape"))

    def _update_progress(self, done, total):
        self.progress["maximum"] = total
        self.progress["value"] = done


# ── Entry point ──────────────────────────────────────────────────────────────

if __name__ == "__main__":
    root = tk.Tk()
    app = ImageScraperApp(root)
    root.mainloop()
